import json
import re
from collections import defaultdict
from datetime import datetime

import openpyxl



# 源文件
read_fileName = 'C:/Users/amos.tong/Desktop/3.5.xlsx'

# 输出文件
out_file_name = 'C:/Users/amos.tong/Desktop/标注数据-3.5.jsonl'

mapping_slots_to_label = {
    "【装修时间": "decorate_time",
    "【装修时间】": "decorate_time",
    "【是否交房】": "has_delivered",
    "【交房时间】": "delivery_time",
    "【量房时间】": "measure_time",
    "【姓氏】": "person",
    "【房屋面积】": "house_area",
    "【房屋类型】": "house_type",
    "【装修类型】": "decorate_type",
    "【装修工程量】": "decorate_content",
    "【房屋装修用途】": "decorate_purpose",
    "【房屋所在城市】": "house_city",
    "【房屋地址】": "house_address",
    "【小区名称】": "community_name"
}

# 创建反向映射字典：value -> key
reverse_mapping = {}
for k, v in mapping_slots_to_label.items():
    # 如果有多个key对应同一个value，我们取第一个遇到的
    if v not in reverse_mapping:
        reverse_mapping[v] = k


def getMappingSlots(slots):
    if slots is None:
        return {}

    # 创建新的slots字典
    new_slots = {}
    for old_key, value in slots.items():
        if old_key in reverse_mapping:
            new_key = reverse_mapping[old_key]
            new_slots[new_key] = value
        else:
            # 如果没有匹配的key，保留原key
            new_slots[old_key] = value
    return new_slots


# 2. 解析send_time（返回datetime对象，无send_time返回None）
def parse_send_time(item):
    if "send_time" not in item:
        return None

    return search_ctime(item["send_time"])


def search_ctime(str_cTime):
    # 提取日期时间部分（如 "2025-06-08 17:13:18"）
    match = re.search(r"(\d{4}-\d{2}-\d{2}).*?(\d{2}:\d{2}:\d{2})", str_cTime)
    if not match:
        return None

    date_str = match.group(1) + " " + match.group(2)
    try:
        return datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
    except:
        return None


def deduplicate(arr):
    seen = {}
    for item in arr:
        # 使用 role + send_time + content 作为唯一键
        send_time = ''
        if "send_time" in item:
            send_time = item["send_time"]
        key = (item["role"], send_time, item["content"])

        # 如果键不存在，直接保存
        if key not in seen:
            seen[key] = item
        else:
            # 如果已存在，检查是否有 slots，优先保留有 slots 的项
            if "slots" in item and "slots" not in seen[key]:
                seen[key] = item
    return list(seen.values())


def merge_two_group(merged_group, current_group):
    """
    合并两条 messages 列表，规则如下：

    - 若长度不等：新数组前半为较短数组，后半补较长数组剩余部分
    - 若长度相等：前 length-1 来自 old_msgs，最后一个来自 new_msgs
    """

    old_msgs = merged_group['messages']
    new_msgs = current_group['messages']

    len_old = len(old_msgs)
    len_new = len(new_msgs)

    if len_old != len_new:
        if len_old < len_new:
            shorter = old_msgs
            longer = new_msgs
            new_msg = shorter + longer[len(shorter):]
            current_group['messages'] = new_msg
        else:
            shorter = new_msgs
            longer = old_msgs
            new_msg = shorter + longer[len(shorter):]
            current_group['messages'] = new_msg

        return current_group
    else:
        if len_old == 0:
            return current_group
        else:
            new_msg = old_msgs[:-1] + [new_msgs[-1]]
            current_group['messages'] = new_msg
            return current_group


def rebuild_result(result_params_arr):
    """
    # result为数组，需要根据result['messages']的length数序排列
    # result排序后，根据result['uid_phone']分组，key为：uid_phone，value取数逻辑示例：
        start_result = result[0]
        second_result = result[1]
        length1= start_result['messages']
        length2= second_result['messages']

        取数逻辑：
            - 新增数组arr，长度和 length1、length2最大长度一致。因为已经排序了，所以默认为length2
            - arr数组的前半部分为两者中短的数组，剩余部分使用长数组减去短数组后的元素填充
            - 如果两个数组长度相等，arr数组的length-1 部分为start_result的length-1，arr数组的最后一个元素为second_result的最后一个元素
            - 返回数组arr
    # 返回分组后的result
    """

    # 先按 messages 长度升序排序
    sorted_result = sorted(result_params_arr, key=lambda x: len(x['messages']))

    grouped = defaultdict(list)

    # 按 uid_phone 分组
    for item in sorted_result:
        grouped[item['uid_phone']].append(item)

    new_result = []

    # 处理每组数据
    for uid_phone, group in grouped.items():
        merged_group = {}

        for i in range(len(group)):
            current_group = group[i]

            if i == 0:
                merged_group = current_group
            else:
                merged_group = merge_two_group(merged_group, current_group)

        # 设置prompt
        # 在merged第一个元素增加prompt
        merged_group['messages'].insert(0, prompt)

        new_result.append(merged_group)

    return new_result


if __name__ == '__main__':
    # 打开文件
    workbook = openpyxl.load_workbook(read_fileName)

    # 选择要操作的工作表
    worksheet = workbook['Sheet1']

    # 获取工作表的行数和列数
    num_rows = worksheet.max_row

    # 存储每一行记录
    result_arr = []

    # 遍历工作表并读取数据
    for row_index in range(2, num_rows + 1):
        try:
            if row_index == 23:
                print(f" ============   over！")

            column2_input = worksheet.cell(row_index, 2).value
            column2_input_array = json.loads(column2_input, strict=False)

            # 电话id uid
            column3_slots = worksheet.cell(row_index, 3).value
            column3_slots_json = json.loads(column3_slots, strict=False)

            column4_output = worksheet.cell(row_index, 4).value
            column4_output_json = json.loads(column4_output, strict=False)
            column4_output_reply = column4_output_json['reply']
            column4_output_slots = column4_output_json['slots']
            column4_output_slots = getMappingSlots(column4_output_slots)

            # 提示语
            column5_prompt = worksheet.cell(row_index, 5).value

            # 大模型返回
            column6_prompt_output = worksheet.cell(row_index, 6).value
            column6_prompt_output_json = json.loads(column6_prompt_output, strict=False)
            column6_prompt_output_answer = column6_prompt_output_json.get('answer')
            # column6_prompt_output_slots = {"decorate_time": "近两三月", "delivery_time": "30天内"}
            column6_prompt_output_slots = column6_prompt_output_json.get("slots")
            column6_prompt_output_slots = getMappingSlots(column6_prompt_output_slots)

            # 时间
            column9_time = worksheet.cell(row_index, 9).value
            cTime = column9_time.strip().replace("cTime：", "")
            cTime_day = search_ctime(cTime)
            if cTime_day is None:
                # cTime_day 解析错误的场景一般是回测数据
                continue

            uid = column3_slots_json.get('chatId')
            phoneId = column3_slots_json.get("phoneId", 0)

            # 如果input的日期和ctime的日期不是同一天，丢弃本消息。（默认为回测数据）
            first_send_time = parse_send_time(column2_input_array[0])
            if cTime_day.date() != first_send_time.date():
                continue

            new_input_messages = []
            new_input_messages.extend(column2_input_array)
            # 拼接当前大模型回复内容
            message = {
                "role": "assistant",
                "send_time": cTime,
                "content": column6_prompt_output_answer,
                "slots": column6_prompt_output_slots,
                "completed_content_gpt": column4_output_reply,
                "completed_slots_gpt": column4_output_slots,
                "completed_content_own": column4_output_reply,
                "completed_slots_own": column4_output_slots,
                "timestamp": 0,
                "edited_time": "",
                "operator": "",
                "has_deleted": False
            }
            new_input_messages.append(message)

            # 提示语
            prompt = {
                "role": "system",
                "content": column5_prompt,
                "timestamp": 0,
                "edited_time": "",
                "operator": "",
                "has_deleted": False,
            }
            result = {
                "uid_phone": f'{uid}%{phoneId}',
                "system_prompt": prompt,
                "uid": uid,
                "phoneId": phoneId,
                "messages": new_input_messages,
                # cTime对应的日期。精确到天
                "cTime_day": cTime_day.strftime("%Y-%m-%d %H:%M:%S")
            }

            result_arr.append(result)


        except Exception as e:
            print(f"{row_index}--{e}")
            continue

    inner_merged_result = rebuild_result(result_arr)
    with open(out_file_name, "w", encoding="utf-8") as f:
        for group in inner_merged_result:
            del group['system_prompt']
            f.write(json.dumps(group, ensure_ascii=False) + "\n")

    # 打印
    # for key, value in inner_merged_result.items():
    #     # if key.endswith("424136515"):
    #     # print(f"key: {json.dumps(value, indent=4, ensure_ascii=False)}")
    #     # print(f"value: {json.dumps(value, indent=4, ensure_ascii=False)}")
    #     print(f"{key}: {json.dumps(value, ensure_ascii=False)}")
