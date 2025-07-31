import ast
import re
import json
from openpyxl import load_workbook


# 大模型提取后的数据
file_1 = 'C:/Users/amos.tong/Desktop/第二批提取优质对话-优质片段0-all-conversation.提取标签(2).xlsx'
# 基于大模型提取的数据，重新人工校准的数据
file_2 = 'C:/Users/amos.tong/Desktop/优质对话-提示词标注.xlsx'

# 最终输出
output_file = 'C:/Users/amos.tong/Desktop/待训练的数据.jsonl'

# 提示语模板
prompt_template = """\
### **角色**
你是土巴兔平台的微信核需客服。当前的时间为：{send_time}

### **历史槽位**
- 装修时间(decorate_time): {decorate_time}
- 是否交房(has_delivered): {has_delivered}
- 交房时间(delivery_time): {delivery_time}
- 量房时间(measure_time): {measure_time}
- 姓氏(person): {person}
- 房屋面积(house_area): {house_area}
- 房屋类型(house_type): {house_type}
- 装修类型(decorate_type): {decorate_type}
- 装修工程量(decorate_content): {decorate_content}
- 房屋装修用途(decorate_purpose): {decorate_purpose}
- 房屋所在城市(house_city): {house_city}
- 房屋地址(house_address): {house_address}
- 小区名称(community_name): {community_name}

### **目标**
收集如下槽位信息：
- 装修时间(decorate_time)
- 是否交房(has_delivered)
- 交房时间(delivery_time)
- 量房时间(measure_time)
- 姓氏(person)
- 房屋面积(house_area)
- 房屋类型(house_type)
- 装修类型(decorate_type)
- 装修工程量(decorate_content)
- 房屋装修用途(decorate_purpose)
- 房屋所在城市(house_city)
- 房屋地址(house_address)
- 小区名称(community_name)

### **默认工作流程**
{default_workflow}\
"""

# 初始化一个空 slots 模板
empty_slots = {
    "decorate_time": "",
    "has_delivered": "",
    "delivery_time": "",
    "measure_time": "",
    "person": "",
    "house_area": "",
    "house_type": "",
    "decorate_content": "",
    "decorate_type": "",
    "decorate_purpose": "",
    "house_city": "",
    "house_address": "",
    "community_name": ""
}

def parse_conversation(text):
    # 初始化消息列表
    messages = []

    # 定义正则表达式匹配模式
    pattern = r'(【(顾问|用户)】：|(顾问|用户)：)(.*)'

    # 按行分割文本
    lines = text.split('\n')

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # 匹配角色和内容
        match = re.match(pattern, line)
        if match:
            # 确定角色
            role = match.group(2) or match.group(3)
            if role == '顾问':
                role = 'assistant'
            else:
                role = 'user'

            # 获取内容
            content = match.group(4).strip()

            # 添加到消息列表
            messages.append({
                "role": role,
                "content": content
            })

    # 构建最终结果
    result = {
        "messages": messages
    }

    return result

def process_excel_file_manual_mark(file_path):
    # 加载Excel文件
    wb = load_workbook(file_path)
    ws = wb.active

    # 准备结果列表
    results = []

    # 遍历每一行（从第二行开始，假设第一行是表头）
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 获取uid和对话记录
        responsible_person = row[0]  # A列 责任人
        uid = row[1]               # B列 uid
        conversation = row[2]      # C列 对话记录
        prompt = row[3]            # D列 prompt（如果有）

        # 解析对话记录
        if conversation:
            parsed = parse_conversation(conversation)
            parsed['uid'] = uid

            # 如果有prompt，添加到消息开头
            if prompt:
                prompt_lines = prompt.split('\n')
                send_time = prompt_lines[0].strip().split("：")[1]
                default_workflow = prompt.replace(prompt_lines[0], '')
                parsed['workflow'] = default_workflow
                parsed['send_time'] = send_time

                parsed['messages'].insert(0, {
                    "role": "system",
                    "content":
                        prompt_template.replace('{send_time}', send_time)
                                          .replace('{default_workflow}', default_workflow)
                })
            results.append(parsed)

    return results


def process_excel_file_mode_mark(file_path):
    # 加载Excel文件
    wb = load_workbook(file_path)
    ws = wb.active

    # 准备结果列表
    results_arr = []

    # 遍历每一行（从第二行开始，假设第一行是表头）
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 获取uid和对话记录
        uid = row[0]               # A列 uid
        conversation_str = row[2]      # C列 messages
        extract_results_str = row[4]   # E列 提取结果_豆包32K


        extract_results = json.loads(extract_results_str, strict=False)


        conversation_arr = ast.literal_eval(conversation_str)

        # 获取 data1 的键并按数字升序排序
        sorted_keys = sorted(extract_results.keys(), key=lambda x: int(x))

        for idx, sorted_key in enumerate(sorted_keys):
            if idx == 0:# 第一个
                for conversation in conversation_arr[:int(sorted_key)]:
                    conversation["slots"] = empty_slots.copy()
                if idx + 1 < len(sorted_keys):# 不是最后一个
                    for conversation in conversation_arr[int(sorted_key):int(sorted_keys[idx + 1])]:
                        conversation["slots"] = extract_results[sorted_key]
            elif idx + 1 < len(sorted_keys):# 中间元素
                for conversation in conversation_arr[int(sorted_key):int(sorted_keys[idx + 1])]:
                    conversation["slots"] = extract_results[sorted_key]
            else: # 最后一个元素
                for conversation in conversation_arr[int(sorted_key):]:
                    conversation["slots"] = extract_results[sorted_key]
        result = {
            "uid": uid,
            "messages": conversation_arr
        }
        results_arr.append(result)
    return results_arr


def filter_parsed_data(data1, data2):
    # 合并手动标记的数据和模型标记的数据。两份数据通过uid进行匹配，并保留手动标记的数据。
    # uid一致，messages中的长度不一致，打印uid，保留手动标记的数据
    # 构建手动标注数据的 UID 到数据的映射
    # 准备最终输出结果列表
    final_results = []
    # 手动和模型 uid匹配不上的数据
    error_uid_arr = []
    mode_data_map = {item["uid"]: item for item in data2}
    # 遍历模型标记的数据，尝试匹配手动标注数据
    for manual_item in data1:
        uid = manual_item["uid"]
        if uid in mode_data_map:
            mode_item = mode_data_map[uid]
            if len(mode_item["messages"]) != len(manual_item["messages"]):
                error_uid_arr.append(uid)
            else:  # 保留手动标注数据
                for i in range(len(manual_item["messages"])):
                    manual_item["messages"][i]["slots"] = mode_item["messages"][i]["slots"]
                    if "send_time" in mode_item["messages"][i]:
                        manual_item["messages"][i]["send_time"] = mode_item["messages"][i]["send_time"]
                # manual_item["send_time"] = mode_item["send_time"]
                # manual_item["slots"] = mode_item["slots"]
                final_results.append(manual_item)
        else:
            error_uid_arr.append(uid)
    return final_results, error_uid_arr


# 使用示例
if __name__ == "__main__":
    # 手动标记的数据
    manual_mark_parsed_data = process_excel_file_manual_mark(file_2)
    # 模型标记的数据
    mode_mark_parsed_data = process_excel_file_mode_mark(file_1)
    # 返回手动标记的数据和丢弃的uid
    final_manual_data, error_uid_arr_return = filter_parsed_data(manual_mark_parsed_data, mode_mark_parsed_data)



    for value in manual_mark_parsed_data:
        print(f" {json.dumps(value, ensure_ascii=False)}")

    print(f"===========================================================================")
    print(f"===========================================================================")
    print(f"===========================================================================")
    for value in mode_mark_parsed_data:
        print(f" {json.dumps(value, ensure_ascii=False)}")
    print(f"@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
    print(f"@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
    for value in final_manual_data:
        print(f"{json.dumps(value, ensure_ascii=False)}")

    print(f'error_uid_arr_return = {error_uid_arr_return}')
    print(f"########################################################################")
    print(f"########################################################################")

    # 输出结果到JSON文件

    with open(output_file, "w", encoding="utf-8") as f:
        for data in final_manual_data:
            f.write(json.dumps(data, ensure_ascii=False) + "\n")