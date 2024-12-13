import requests
import openpyxl

'''
    批量验证应答V2意图
'''
def call_nerAndcategoryV2(question, content):
    url = "http://192.168.41.112:40015/nlp/nerAndcategoryV2"
    data = {
        "bussinessKey": "intelligentPlatform",
        "bussinessId": 123456,
        "question": question,
        "content": content,
        "currentTime": 1724898903
    }
    res = requests.post(url=url, json=data)

    return res.json()

def call_QA(content):
    url = "http://192.168.41.112:40015/nlp/qaMulity"
    data = {
        "bussinessKey": "intelligentPlatform",
        "bussinessId": 123456789,
        "sourceType": "WX_CS",
        "content": content,
        "version": "1.0"
    }
    res = requests.post(url=url, json=data)

    return res.json()

if __name__ == '__main__':
    # 打开文件
    # fileName = 'C:\\Users\\amos.tong\\Desktop\\\DBeaver_export\\_应答问到了量房时间_完整语料_并且需要标记转人工位置_模板表格.xlsx'
    fileName = 'C:\\Users\\amos.tong\\Desktop\\\DBeaver_export\\2024-11-25-顾问+用户一问一答.xlsx'
    workbook = openpyxl.load_workbook(fileName)

    # 获取工作表列表
    sheet_names = workbook.get_sheet_names()
    print("工作表列表:", sheet_names)

    # 选择要操作的工作表
    #worksheet = workbook['Sheet7']
    worksheet = workbook['2024-11-25 顾问+用户']

    # 获取工作表的行数和列数
    num_rows = worksheet.max_row
    num_cols = worksheet.max_column
    print(f"工作表行数: {num_rows} 工作表列数: {num_cols}")

    print("行号&顾问提问&用户回复&文本分类结果&QA识别结果")
    # 遍历工作表并读取数据
    for row_index in range(10510, num_rows+1):
        col_robot = worksheet.cell(row_index, 1).value
        user_robot = worksheet.cell(row_index, 2).value
        data = call_nerAndcategoryV2(col_robot, user_robot)
        data_qa = call_QA(user_robot)
        print(f"{row_index}&{col_robot}&{user_robot}&{data['data']['categoryInfoList']}&{data_qa['data']['qaMulityVO']}")

    # 关闭工作簿
    workbook.close()
    print(" ============   over！")