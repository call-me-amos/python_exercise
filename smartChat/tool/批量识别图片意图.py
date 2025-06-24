import json

import openpyxl
import requests

'''
    自定义批量执行URL
'''


# 企微表情包识别
urlForQiweiEmoji = "http://192.168.11.116:8970/cls_v2/pic"

def call_http(param_image_url, messageType):
    if messageType == 3:
        data = {
            "img_url": param_image_url + "?x-oss-process=image/format,jpg/quality,95/resize,w_1200,h_1200",
        }
    else:
        data = {
            "img_url": param_image_url,
        }
    res = requests.post(url=urlForQiweiEmoji, json=data)
    try:
        return 200, res.json()
    except requests.exceptions.RequestException as e:
        # 捕获所有requests库抛出的异常
        return 500, str(e)  # 返回异常信息作为字符串
    except ValueError:
        # 如果res.json()失败（例如，因为响应不是有效的JSON），这个异常会被捕获
        return 500, res.text  # 返回响应文本



if __name__ == '__main__':
    # 打开文件
    fileName = 'C:\\Users\\amos.tong\\\Desktop\\\DBeaver_export\\语料-图片+企微表情包.xlsx'
    workbook = openpyxl.load_workbook(fileName)

    # 获取工作表列表
    sheet_names = workbook.get_sheet_names()

    # 选择要操作的工作表
    # worksheet = workbook['企微表情']
    worksheet = workbook['图片']

    # 获取工作表的行数和列数
    num_rows = worksheet.max_row
    num_cols = worksheet.max_column
    print(f"工作表行数: {num_rows} 工作表列数: {num_cols}")

    # 遍历工作表并读取数据
    for row_index in range(2, num_rows+1):
        if row_index > 1000:
            print(f" ============   over！")
            break
        img_url_json_str = worksheet.cell(row_index, 9).value
        img_url_json = json.loads(img_url_json_str)
        url = img_url_json["url"]
        code, result = call_http(url, worksheet.cell(row_index, 8).value)
        if code == 200:
            if result['code'] == 200:
                print(f"success  {row_index}  {url}  {result['data']['cls']}")
            else:
                print(f"fail  {row_index}  {url}  {json.dumps(result)}")
        else:
            print(f"error  {row_index}  {url}  {result}")




