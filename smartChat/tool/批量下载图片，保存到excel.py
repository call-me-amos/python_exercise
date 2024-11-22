import openpyxl
import requests
from openpyxl.drawing.image import Image
from io import BytesIO
from openpyxl.utils import get_column_letter

def download_image(url):
    response = requests.get(url)
    response.raise_for_status()  # Raise an HTTPError for bad responses (4xx and 5xx)
    return BytesIO(response.content)

if __name__ == "__main__":
    # 打开文件
    fileName = 'C:\\Users\\amos.tong\\Desktop\\\表情包2\\\表情包URL.xlsx'
    workbook = openpyxl.load_workbook(fileName)

    # 获取工作表列表
    sheet_names = workbook.get_sheet_names()
    print("工作表列表:", sheet_names)

    # 选择要操作的工作表
    worksheet = workbook['Sheet1']
    # 需要写入的sheet
    ws = workbook.active
    ws.title = "Sheet2"

    # 获取工作表的行数和列数
    num_rows = worksheet.max_row
    num_cols = worksheet.max_column
    print(f"工作表行数: {num_rows} 工作表列数: {num_cols}")

    # 遍历工作表并读取数据
    for row_index in range(2, num_rows+1):
        img_url = worksheet.cell(row_index, 1).value
        image_data = download_image(img_url)
        img = Image(image_data)
        output_row = row_index
        ws[f'A{output_row}'] = img_url
        ws.add_image(img, f'B{row_index}')

    fileName = 'C:\\Users\\amos.tong\\Desktop\\\表情包2\\\表情包URL-out.xlsx'
    workbook.save(fileName)
    print(" ============   over！")