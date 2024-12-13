import requests
import openpyxl
import mimetypes
import os
from urllib.parse import urlparse
import json

def get_image_extension(url):
    # Send a HEAD request to get the Content-Type
    response = requests.head(url, allow_redirects=True)
    content_type = response.headers.get('Content-Type')

    if not content_type:
        raise ValueError("Could not determine Content-Type from URL")

    # Guess the extension based on the Content-Type
    extension = mimetypes.guess_extension(content_type)

    # Fallback: if no extension is guessed, use the URL's path extension
    if not extension:
        parsed_url = urlparse(url)
        file_extension = os.path.splitext(parsed_url.path)
        if file_extension:
            extension = file_extension[1:]  # Remove the leading dot

    return extension

def download_image(url, output_path):
    response = requests.get(url, stream=True)
    response.raise_for_status()

    # Get the guessed extension
    extension = get_image_extension(url)
    if not extension:
        extension = '.bin'  # Fallback to a generic extension if none can be guessed

    # Ensure the output path has the correct extension
    output_filename = os.path.basename(output_path)
    if not output_filename.endswith(extension):
        if isinstance(extension, tuple):
            output_path = os.path.splitext(output_path)[0] + str(extension[0])
        else:
            output_path = os.path.splitext(output_path)[0] + extension

    with open(output_path, 'wb') as f:
        for chunk in response.iter_content(chunk_size=8192):
            f.write(chunk)

    print(f"Downloaded {url} to {output_path}")


if __name__ == '__main__':
    # 打开文件
    fileName = 'C:\\Users\\amos.tong\\Desktop\\\DBeaver_export\\\语料-图片+企微表情包.xlsx'
    workbook = openpyxl.load_workbook(fileName)

    # 获取工作表列表
    sheet_names = workbook.get_sheet_names()

    # 选择要操作的工作表
    # worksheet = workbook['企微表情']
    worksheet = workbook['企微表情']

    # 获取工作表的行数和列数
    num_rows = worksheet.max_row
    num_cols = worksheet.max_column
    print(f"工作表行数: {num_rows} 工作表列数: {num_cols}")

    # 遍历工作表并读取数据
    for row_index in range(2, num_rows+1):
        if row_index > 1000:
            print(f" 大小超额了 ============   over！")
            break
        img_url_json_str = worksheet.cell(row_index, 9).value
        img_url_json = json.loads(img_url_json_str)
        url = img_url_json["url"]
        output_path = f"C:\\Users\\amos.tong\\Desktop\\\表情包回测\\企微表情\\企微表情_{row_index}.png"
        download_image(url, output_path)

    # 关闭工作簿
    workbook.close()
    print(" ============   over！")