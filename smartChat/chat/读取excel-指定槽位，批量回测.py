import time
import requests
import uuid
import openpyxl


def test_bot(user, text):
    url = "http://10.4.42.48:40121/tls/smartChat/testReply"
    mode = 1
    if text == "":
        print("开启会话: chatId:", user)
        mode = 2
    dataJson = {
        "user": user,
        "weChat": "tongzhiwei",
        "eventType": mode,
        "preRobotAskId": 119,  # 提问槽位
        "messageId": uuid.uuid4().hex,
        "reply": text

    }
    res = requests.post(url=url, json=dataJson)
    return res.json()


if __name__ == '__main__':
    # 打开文件
    workbook = openpyxl.load_workbook('C:\\Users\\amos.tong\\Desktop\\历史记录-标签提取\\标签提取结果-20231024.xlsx')

    # 获取工作表列表
    sheet_names = workbook.get_sheet_names()
    print("工作表列表:", sheet_names)

    # 选择要操作的工作表
    worksheet = workbook['用户回复-房屋类型']  # 替换为你要读取的工作表名称

    # 获取工作表的行数和列数
    num_rows = worksheet.max_row
    num_cols = worksheet.max_column
    print(f"工作表行数: {num_rows}")
    print(f"工作表列数: {num_cols}")

    user_name = ''
    # 遍历工作表并读取数据
    for row_index in range(1, num_rows+1):
        for col_index in range(1, num_cols+1):
            # 单元格从（1,1） 开始
            cell_value = worksheet.cell(row_index, col_index).value
            print(f"行{row_index}, 列{col_index}: {cell_value}")
            if str(cell_value) == '':
                print(f"行{row_index}, 列{col_index}: 空字符串")
                continue
            user_name = str(int(time.time())) + "_" + uuid.uuid4().hex
            test_bot(user_name, '')
            result = test_bot(user_name, str(cell_value))
            print(result)
    # 关闭工作簿
    workbook.close()
    print(" ============   over！")
