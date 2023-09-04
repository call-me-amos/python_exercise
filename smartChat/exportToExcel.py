from logging import exception

import openpyxl
import json
import requests
import time
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

uid = 13808
erp_ticket = "eds5uqLStVvQCjKEveag3aV3sRpwfEDTa8ZocxRUtT3v_VMt_YlocfJG-sxonMVIdsj9e8Se2B7kJKLHszicTBd_pzDOaaV1sQVS5CXK2wuKspIyQD_4lwgyzEAJHTpq"

# 查询群聊列表
query_session = "https://apigw.to8to.com/cgi/erp/fls/imServiceApi/record/querySessionPage?uid=" + str(uid) \
                + "&appName=erp&tu-customize-frontId=11ac17a0-c1ac-11ed-9490-73c6c5846b39&refsrc=%2Fim%2Fchat-history" \
                  "&ticket=" + erp_ticket

# 查询详细聊天记录
query_record = "https://apigw.to8to.com/cgi/erp/fls/imServiceApi/record/queryByConOrTime?uid=790761&appName=erp&tu" \
               "-customize-frontId=11ac17a0-c1ac-11ed-9490-73c6c5846b39&refsrc=%2Fim%2Fchat-history&ticket=" + \
               erp_ticket

headers = {
    'Authorization': 'Basic ZGV2OnRlc3Q=',
    'Connection': 'keep-alive',
    'accept': 'application/json, text/plain, */*',
    'content-type': 'application/x-www-form-urlencoded;charset=UTF-8',
    'origin': 'https://scm.to8to.com',
    'referer': 'https://scm.to8to.com',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 '
                  'Safari/537.36'
}


# 查询会话列表
def querySession(start_page, start_time, end_time):
    args = {
        "args": {
            "search": {
                "createTime_gte": start_time,
                "createTime_lte": end_time,
                "groupName": '在线出图APP',
                "targetType": 3
            },
            "page": start_page,
            "size": 100
        }
    }
    response = requests.post(query_session, headers=headers, data=json.dumps(args), verify=False)
    resp_json = json.loads(response.text)
    status = resp_json['status']
    if status != 200:
        print("请求异常：" + str(resp_json))
    return resp_json


# 查找详细聊天记录
def queryChatRecord(session_id):
    args = {
        "args": {
            "page": 1,
            "size": 300,
            "sessionId": session_id
        }
    }
    response = requests.post(query_record, headers=headers, data=json.dumps(args), verify=False)
    resp_json = json.loads(response.text)
    status = resp_json['status']
    if status != 200:
        print(resp_json)
    return resp_json


if __name__ == '__main__':
    # 获取 工作簿对象
    workbook = openpyxl.load_workbook('csv/聊天记录.xlsx')

    # 获取表对象
    worksheet = workbook['Sheet1']

    # 开始和结束时间
    startTime = 1679328000
    endTime = 1680001800

    for page in range(10000):
        print("page=" + str(page))
        session_result = querySession(page + 1, startTime, endTime)
        session_rows = session_result['result']['rows']
        if len(session_rows) <= 0:
            break
        for i in range(len(session_rows)):
            group = session_rows[i]
            group_id = group['sessionId']
            memberName = group['memberName']
            groupName = group['groupName']
            if len(memberName) < 1:
                print("异常的会话 group_id=" + str(group_id))
                continue

            record = queryChatRecord(group_id)
            data = record['result']['rows']
            total = record['result']['total']
            if total < 1:
                continue

            # 插入表头
            worksheet.cell(row=1, column=(i + 1) + (page * 100)).value = str(group_id) + '---' + groupName
            for j in range(len(data)):
                fromUserName = ''
                if 'fromUserName' in data[j]:
                    fromUserName = data[j]['fromUserName']
                else:
                    fromUserName = data[j]['fromUserId']
                sendTime = data[j]['sendTime']
                sendTimeStr = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(int(sendTime) / 1000))
                title = fromUserName + ' ' + sendTimeStr
                print(title)
                worksheet.cell(row=2 * (j + 1), column=(i + 1) + (page * 100)).value = title
                contentObj = json.loads(data[j]['content'])
                content = ''
                if contentObj.__contains__('bizObj'):
                    bizObj = contentObj['bizObj']
                    if bizObj.__contains__('content'):
                        content = bizObj['content']
                    elif bizObj.__contains__('coverUrl'):
                        content = '图片' + bizObj['coverUrl']
                elif contentObj.__contains__('content'):
                    content = contentObj['content']
                print(content)
                worksheet.cell(row=2 * (j + 1) + 1, column=(i + 1) + (page * 100)).value = content
            print('===================================== ' + groupName)
    workbook.save(filename='csv/聊天记录-2023-03-29.xlsx')
print("----------   over --------------")